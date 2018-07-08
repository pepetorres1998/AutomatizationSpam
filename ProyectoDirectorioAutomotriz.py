#IMPORTAR MODULOS
from bs4 import BeautifulSoup as bf
import requests as rqs
import re
import smtplib as sm
import openpyxl as pxl
import time
import sys
from datetime import datetime

#CONSTANTS
list_empty = []


header_rqs = {
	'User-Agent': 'JOTGA Technologies',
	'From': 'jotgatech@gmail.com'
}
control = (int(sys.argv[2])-1)*25

#UTIL FUNCTIONS
def index_page(a):
	"""
		Return '' if page is 0, return page index number otherwise
	"""
	if(a>0):
		return str(a)
	else:
		return ''

def still_num(r, a):
	"""
		Counter for the number of int in a string
	"""
	while True:
		try:
			int(r[a+1])
			a = a+1
			still_num(r, a)
		except ValueError:
			a = a+1
			return a

def get_links(index_links, index_pattern, links_nombres, links_href):
	"""
		Gets the links in the index.
	"""
	for i in index_links:
		if(re.search(index_pattern, str(i.prettify()))):
			temp = re.search(index_pattern, str(i.prettify()))
			temp1 = str(i)
			links_href.append(temp1[temp.start():still_num(temp1, temp.end())])
			links_nombres.append(i.text)

def use_links(links_href, c_index, index_noemail, noemail_names, index_datos, links_nombres, tablas_company, tablas_nombres, tablas_puesto, tablas_correo, tablas_tel):
	"""
	Use the links from get links
	"""
	c_pags = 0

	for i in range(0, len(links_href), 2):
		print("\tIniciando pagina "+str(c_index+(c_pags+1)))
		pagina_ = rqs.get(links_href[i], headers = header_rqs)
		pagina_soup = bf(pagina_.content, 'lxml')
		print('\t\t'+links_href[i])
		print("\t\tSTATUS_Pagina: "+str(pagina_.status_code))
		print("\t\tSleeping")
		print('\t\t'+str(datetime.now().time()))
		time.sleep(5)
		pagina_datos = pagina_soup.find_all('div', class_='col-md-6')
		pagina_datos = pagina_datos[1].find('p')
		try:
			if pagina_soup.find_all('td') == list_empty:
				pagina_tabla = []
				raise IndexError
			else:
				pagina_tabla = pagina_soup.find_all('td')
				print('\t\t\tTABLA ENCONTRADA')
				get_table(pagina_tabla, i+1, links_nombres, tablas_company, tablas_nombres, tablas_puesto, tablas_correo, tablas_tel)
				#print(tablas_company)	#CORRECTO
				#print(tablas_nombres)	#CORRECTO
		except IndexError:
			print("\t\t\tNO HAY TABLA")
			index_noemail.append(pagina_datos)
			noemail_names.append(i)
		index_datos.append(pagina_datos)

		print('\t\t\tDATOS ENCONTRADOS')

		c_pags = c_pags+1

def get_table(pagina_tabla, c_empresa, links_nombres, tablas_company, tablas_nombres, tablas_puesto, tablas_correo, tablas_tel):
	"""
	Function that gets the table data from the web page
	"""
	for k in range(len(pagina_tabla)):
		if(k%4==0):
			tablas_company.append(links_nombres[c_empresa])
			tablas_nombres.append(pagina_tabla[k].text)
		elif(k%4==1):
			tablas_puesto.append(pagina_tabla[k].text)
		elif(k%4==2):
			tablas_correo.append(pagina_tabla[k].text)
		elif(k%4==3):
			tablas_tel.append(pagina_tabla[k].text)

def do_excel(tablas_company, tablas_nombres, tablas_puesto, tablas_correo, tablas_tel, c_index):
	"""
		Do the .xlsx file
	"""
	wb = pxl.Workbook()
	ws = wb.active
	ws.title = "Index "+str((c_index/25)+1)
	max_lenght = 0
	#CONSTANTES
	firstFill = pxl.styles.fills.PatternFill(start_color = '00FF0000', end_color = '00FF0000', fill_type = 'solid')
	secondFill = pxl.styles.fills.PatternFill(start_color = '00FFFF00', end_color = '00FFFF00', fill_type = 'solid')
	thirdFill = pxl.styles.fills.PatternFill(start_color = '00FF00FF', end_color = '00FF00FF', fill_type = 'solid')
	fourthFill = pxl.styles.fills.PatternFill(start_color = '0000FF00', end_color = '0000FF00', fill_type = 'solid')
	fifthFill = pxl.styles.fills.PatternFill(start_color = '0000FFFF', end_color = '0000FFFF', fill_type = 'solid')
	sixthFill = pxl.styles.fills.PatternFill(start_color = '000000FF', end_color = '000000FF', fill_type = 'solid')
	Fills = [secondFill, thirdFill, fourthFill, fifthFill, sixthFill]
	centerAlignment = pxl.styles.Alignment(horizontal='center', vertical='center', wrapText=True)
	#centerAlignment = pxl.styles.Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
	ex_encabezado = ['Empresa', 'Nombre', 'Puesto', 'Correo', 'Telefono']
	#FOR DE ENCABEZADO
	for i in range(1, 2):
		for j in range(1, 6):
			current = ws.cell(row = i, column = j, value = ex_encabezado[j-1])
			current.alignment = centerAlignment
			current.font = pxl.styles.fonts.Font(bold=True)
			current.fill = firstFill

	#FOR CUERPO
	for row in range(2, len(tablas_nombres)+2):
		for col in range(1, len(ex_encabezado)+1):
			if(col == 1):
				current = ws.cell(row=row, column=col, value=tablas_company[row-2])
			elif(col == 2):
				current = ws.cell(row=row, column=col, value=tablas_nombres[row-2])
			elif(col == 3):
				current = ws.cell(row=row, column=col, value=tablas_puesto[row-2])
			elif(col == 4):
				current = ws.cell(row=row, column=col, value=tablas_correo[row-2])
			elif(col == 5):
				current = ws.cell(row=row, column=col, value=tablas_tel[row-2])
			current.alignment = centerAlignment
			if(tablas_company[row-2] != tablas_company[row-3]):
				firstFill = Fills[((row-2)%5) if current.fill != Fills[(row-2)%5] else (((row-2)%5)+1)]
				current.fill = Fills[((row-2)%5) if current.fill != Fills[(row-2)%5] else (((row-2)%5)+1)]
			else:
				current.fill = firstFill
	try:
		wb.save('Index '+str((c_index/25)+1)+'.xlsx')
		print('\t\t\tFichero guardado')
	except:
		print('\t\t\tERROR AL GUARDAR')

def send_mail():
	"""
		Function that sends the emails
	"""
	mail_remitente = "jotgatech@gmail.com"
	mail_mensaje = """
<br/><br/>
<ul style="list-style-type:none">
	<li style="font-size:125%">Llega a m&aacute;s de 4000 clientes.
	<li style="font-size:125%">Obt&eacute;n los datos de m&aacute;s de 4000 empresas.
	<li style="font-size:125%">Implementa la tecnolog&iacute;a de una manera inteligente.
</ul>
<p>Visita este link para mejor info: <a href='https://imgur.com/a/FAM2dZK'>Imgur</a>
</p>
<p style="font-size:80%">Este correo fue creado y enviado por JOTGA Technologies, a trav&eacute;s de Python.
Para informaci&oacute;n o comentarios: jotgatech@gmail.com</p>
"""
	mail_asunto = "Servicio automatizacion de correos."
	server = sm.SMTP_SSL('smtp.gmail.com:465')
	server.ehlo()
	server.login(mail_remitente, 'panteras_lynx98') #CORRIGE CONTRASEÑA
	mail_destinatario = tablas_correo
	mail_email = """From: <%s>
To: <%s>
MIME-Version: 1.0
Content-type: text/html
Subject: %s


%s
""" % (mail_remitente, mail_destinatario, mail_asunto, mail_mensaje)
	server.sendmail(mail_remitente, mail_destinatario, mail_email)
	server.close()

def send_mail_try():
	try:
		send_mail()
		print('\t\t\tMail sended')
	except sm.SMTPRecipientsRefused:
		print('\t\t\tNO HAY EMAIL')

def save_data(index_datos, c_index, index_noemail, links_nombres, noemail_names):
	"""
		Function that saves data in .txt doc
	"""
	with open('Datos Index '+str((c_index/25)+1)+'.txt', 'w') as fila_datos:
		for i in range(len(index_datos)):
			fila_datos.write(links_nombres[(i*2)+1])
			fila_datos.write('\n')
			fila_datos.write(index_datos[i].text)
			fila_datos.write('\n')
		print('\t\t\tDATOS GUARDADOS')
	with open('Index no email '+str((c_index/25)+1)+'.txt', 'w') as fila_noemail:
		for i in range(len(index_noemail)):
			fila_noemail.write(links_nombres[noemail_names[i]+1]+'\n')
			fila_noemail.write(index_noemail[i].text+'\n')

#PROGRAMA MAIN
def main():
	"""
		main of program
	"""
	c_index = (int(sys.argv[1])-1)*25
	c_correos = 0
	while True:
		tablas_company = []
		tablas_nombres = []
		tablas_puesto = []
		tablas_correo = []
		tablas_tel = []
		index_datos = []
		index_noemail = []
		noemail_names = []
		print("Iniciando index "+str((c_index/25)+1))
		index_ = rqs.get('http://www.directorioautomotriz.com.mx/core/Busqueda/Libre/Index/'+index_page(c_index)+'/', headers = header_rqs)
		print("Status: "+str(index_.status_code))
		print(index_.url)
		print(str(datetime.now().time()))
		print("Sleeping")
		time.sleep(5)
		index_soup = bf(index_.content, 'html.parser')
		#print(index_soup.prettify())#CORRECTO
		index_links = index_soup.find_all('a')
		#print(index_links)
		index_pattern = r'http://www.directorioautomotriz.com.mx/core/Negocio/detail/'
		links_nombres = []
		links_href = []
		get_links(index_links, index_pattern, links_nombres, links_href)
		print(len(links_href))
		#print(len(links_nombres))
		#print("\n".join(links_href))
		#print("SEPARADOR\n\n\n")			#Correcto
		#print("\n".join(links_nombres))
		#print(links_nombres)
		use_links(links_href, c_index, index_noemail, noemail_names, index_datos, links_nombres, tablas_company, tablas_nombres, tablas_puesto, tablas_correo, tablas_tel)
		#FUNCION EXCEL
		if(len(tablas_company)>0):
			do_excel(tablas_company, tablas_nombres, tablas_puesto, tablas_correo, tablas_tel, c_index)
		c_correos = c_correos + len(tablas_correo)
		#FUNCION EMAILS
		#print(tablas_correo)
		#send_mail_try()
		print('\t\t\t'+str(datetime.now().time()))
		save_data(index_datos, c_index, index_noemail, links_nombres, noemail_names)
		if(c_index == control):
			print('FINALIZANDO PROGRAMA')
			break
		else:
			print('CONTINUANDO')
			c_index = c_index+25
			continue

	print("Numero de correos enviados: "+str(c_correos))

if(__name__=='__main__'):
	main()
